"""
FASE 3 — Servidor local FastAPI
Correr: python servidor.py
Abre automáticamente en http://localhost:8000
"""

import asyncio, json, os, shutil, tempfile, webbrowser
from datetime import datetime
from threading import Timer

import uvicorn
from fastapi import FastAPI, File, Form, UploadFile, Request
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests as req_lib
from hardgamers_scraper import scrape_hardgamers
from competidores_scraper import scrape_competidores

app = FastAPI()
TEMP_DIR = tempfile.mkdtemp()


def obtener_dolar():
    try:
        r = req_lib.get("https://dolarapi.com/v1/dolares/oficial", timeout=5)
        return float(r.json()["venta"])
    except Exception:
        return None


def leer_productos_calcular(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".txt":
        productos = []
        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                nombre = line.strip()
                if not nombre or nombre.startswith("#"):
                    continue
                if "|" in nombre:
                    partes = [p.strip() for p in nombre.split("|")]
                    nombre = partes[0]
                    ii_raw = partes[1].upper() if len(partes) > 1 else "NO"
                    imp_int = 10.5 if ii_raw in ("SI", "SÍ", "S", "1", "X") else 0.0
                else:
                    imp_int = 0.0
                if nombre:
                    productos.append({"nombre": nombre, "imp_int": imp_int})
        return productos
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    productos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0]
        if val and str(val).strip():
            imp_int = str(row[1]).strip().upper() if len(row) > 1 and row[1] else "NO"
            tiene_imp_int = imp_int in ("SI", "SÍ", "S", "1", "TRUE", "X")
            productos.append({
                "nombre": str(val).strip(),
                "imp_int": 10.5 if tiene_imp_int else 0.0
            })
    return productos

def leer_productos_excel(path):
    return leer_productos_calcular(path)


def calcular_precio_venta(precio_comp, pct_menos):
    return precio_comp * (1 - pct_menos / 100)


def calcular_precio_compra_usd(precio_venta, margen, iva, imp_int, dolar):
    total_impuestos = iva / 100 + imp_int / 100
    precio_neto_ars = (precio_venta / (1 + margen / 100)) / (1 + total_impuestos)
    return precio_neto_ars / dolar


def filtrar_y_ordenar(resultados, tiendas_excluir):
    excluir = [t.lower().strip() for t in tiendas_excluir if t.strip()]
    filtrados = [r for r in resultados if r["precio"] and r["tienda"].lower() not in excluir]
    return sorted(filtrados, key=lambda x: x["precio"])


def recalcular_filas(filas, margen, iva, dolar):
    """Recalcula precio_compra_usd para todas las filas con el margen/iva/dolar actuales."""
    resultado = []
    for f in filas:
        f2 = dict(f)
        if f2.get("precio_venta") and dolar:
            imp_int = f2.get("imp_int", 0) or 0
            total_imp = iva / 100 + imp_int / 100
            neto = (f2["precio_venta"] / (1 + margen / 100)) / (1 + total_imp)
            f2["precio_compra_usd"] = round(neto / dolar, 2)
        resultado.append(f2)
    return resultado


def exportar_excel_vendedor_simple(filas, path):
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Para Vendedor"
    for col, (h, w) in enumerate(zip(["Producto", "Precio Compra USD"], [50, 20]), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = PatternFill("solid", start_color="1F3864")
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, fila in enumerate(filas, 2):
        c1 = ws.cell(row=i, column=1, value=fila["producto"])
        c1.font = Font(name="Arial", size=10)
        c1.border = border
        c2 = ws.cell(row=i, column=2, value=fila["precio_compra_usd"])
        c2.font = Font(name="Arial", size=10, bold=True)
        c2.number_format = '"U$D "#,##0.00'
        c2.alignment = center
        c2.border = border
    ws.freeze_panes = "A2"
    wb.save(path)


def exportar_excel_analisis(filas, path, dolar, margen, iva):
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    d_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análisis"
    headers    = ["Producto", "Precio Venta", "Competidor", "Tienda", "Pos.", "% Menos", "Precio Compra USD", "Dólar", "IVA%", "Imp. Interno", "Margen%"]
    col_widths = [45, 15, 45, 20, 6, 8, 18, 14, 8, 12, 9]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = PatternFill("solid", start_color="1F3864")
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, fila in enumerate(filas, 2):
        bg = "FFFFFF" if i % 2 == 0 else "F2F2F2"
        imp_int_label = "✓ 10.5%" if fila.get("imp_int", 0) > 0 else "No"
        vals = [
            fila["producto"],
            fila.get("precio_venta"),
            fila.get("competidor_nombre", ""),
            fila.get("competidor_tienda", ""),
            fila.get("posicion_real"),
            fila.get("pct_menos"),
            fila.get("precio_compra_usd"),
            dolar,
            iva,
            imp_int_label,
            margen,
        ]
        fmts = [None, '"$"#,##0', None, None, None, '0.0"%"', '"U$D "#,##0.00', '"$"#,##0', '0.0"%"', None, '0.0"%"']
        for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = d_font
            c.fill = PatternFill("solid", start_color=bg)
            c.border = border
            if fmt:
                c.number_format = fmt
                c.alignment = center
            elif col == 10:
                c.alignment = center
                c.font = Font(name="Arial", size=10,
                              bold=fila.get("imp_int", 0) > 0,
                              color="00AA44" if fila.get("imp_int", 0) > 0 else "999999")
    ws.freeze_panes = "A2"
    wb.save(path)


def exportar_excels(filas, dolar, config, nombre_base):
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    path_analisis = os.path.join(TEMP_DIR, f"{nombre_base}_analisis_{fecha}.xlsx")
    path_vendedor = os.path.join(TEMP_DIR, f"{nombre_base}_vendedor_{fecha}.xlsx")

    exportar_excel_analisis(filas, path_analisis, dolar, config["margen"], config["iva"])

    # Vendedor solo con filas HG (sin compragamer)
    filas_hg = [f for f in filas if not f.get("es_compragamer")]
    exportar_excel_vendedor_simple(filas_hg, path_vendedor)

    return path_analisis, path_vendedor


def leer_productos_comparar(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".txt":
        productos = []
        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                nombre = line.strip()
                if not nombre or nombre.startswith("#"):
                    continue
                if "|" in nombre:
                    partes = [p.strip() for p in nombre.split("|")]
                    nombre = partes[0]
                    try:
                        precio_usd = float(partes[1]) if len(partes) > 1 and partes[1] else 0.0
                    except:
                        precio_usd = 0.0
                    try:
                        iva = float(str(partes[2]).replace("%","")) if len(partes) > 2 and partes[2] else 21.0
                    except:
                        iva = 21.0
                    ii_raw = partes[3].upper() if len(partes) > 3 else "NO"
                    imp_int = 10.5 if ii_raw in ("SI", "SÍ", "S", "1", "X") else 0.0
                else:
                    precio_usd = 0.0
                    iva = 21.0
                    imp_int = 0.0
                if nombre:
                    productos.append({"nombre": nombre, "precio_usd": precio_usd, "iva": iva, "imp_int": imp_int})
        return productos
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    productos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue
        nombre     = str(row[0]).strip()
        precio_usd = float(row[1]) if len(row) > 1 and row[1] else 0.0
        iva_raw = str(row[2]).strip().replace("%", "") if len(row) > 2 and row[2] else "21"
        try:
            iva = float(iva_raw)
        except:
            iva = 21.0
        imp_raw = str(row[3]).strip().upper() if len(row) > 3 and row[3] else "NO"
        imp_int = 10.5 if imp_raw in ("SI", "SÍ", "S", "1", "TRUE", "X") else 0.0
        productos.append({"nombre": nombre, "precio_usd": precio_usd, "iva": iva, "imp_int": imp_int})
    return productos


@app.post("/api/comparar_stream")
async def comparar_stream(
    archivo: UploadFile = File(...),
    margen: float = Form(20.0),
    dolar: float = Form(...),
    tiendas_cmp: str = Form(""),
):
    ext = os.path.splitext(archivo.filename)[1].lower() or ".xlsx"
    path_excel = os.path.join(TEMP_DIR, "comparar" + ext)
    with open(path_excel, "wb") as f:
        shutil.copyfileobj(archivo.file, f)

    productos = leer_productos_comparar(path_excel)
    async def generar():
        total = len(productos)
        for i, prod in enumerate(productos, 1):
            nombre     = prod["nombre"]
            precio_usd = prod["precio_usd"]
            iva        = prod["iva"]
            imp_int    = prod["imp_int"]

            msg_prog = json.dumps({"tipo": "progreso", "actual": i, "total": total, "producto": nombre})
            yield "data: " + msg_prog + "\n\n"

            total_imp = iva / 100 + imp_int / 100
            mi_precio = precio_usd * dolar * (1 + margen / 100) * (1 + total_imp)

            resultados_hg = await scrape_hardgamers(nombre, max_results=50)
            resultados_cg = await scrape_competidores(nombre)
            todos = resultados_hg + resultados_cg
            ordenados = sorted([r for r in todos if r.get("precio")], key=lambda x: x["precio"])

            # Agrupar por tienda — mejor precio de cada una
            por_tienda = {}
            for r in ordenados:
                tnombre = r["tienda"].strip()
                if tnombre not in por_tienda:
                    por_tienda[tnombre] = {"nombre": r["nombre"], "precio": r["precio"],
                                           "tienda": tnombre, "url": r.get("url", "")}

            # Estado genérico (más barato del mercado)
            mercado_precio = ordenados[0]["precio"] if ordenados else None
            mercado_tienda = ordenados[0]["tienda"] if ordenados else None
            mercado_nombre = ordenados[0]["nombre"] if ordenados else None

            if mercado_precio:
                diff_pct = (mi_precio - mercado_precio) / mercado_precio * 100
                if diff_pct < -5:
                    estado = "barato"
                elif diff_pct > 5:
                    estado = "caro"
                else:
                    estado = "par"
            else:
                estado = "sin_datos"

            fila = {
                "producto": nombre,
                "precio_usd": precio_usd,
                "mi_precio": round(mi_precio, 2),
                "mercado_precio": round(mercado_precio, 2) if mercado_precio else None,
                "mercado_tienda": mercado_tienda,
                "mercado_nombre": mercado_nombre,
                "por_tienda": por_tienda,
                "estado": estado,
                "iva": iva,
                "imp_int": imp_int,
            }
            msg_res = json.dumps({"tipo": "resultado", "fila": fila})
            yield "data: " + msg_res + "\n\n"

        msg_fin = json.dumps({"tipo": "fin", "dolar": dolar})
        yield "data: " + msg_fin + "\n\n"

    return StreamingResponse(generar(), media_type="text/event-stream")


@app.post("/api/salir")
async def salir():
    import threading
    def _stop():
        import time, os, signal
        time.sleep(0.5)
        os.kill(os.getpid(), signal.SIGTERM)
    threading.Thread(target=_stop, daemon=True).start()
    return {"ok": True}


@app.get("/", response_class=HTMLResponse)
async def index():
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "web.html"), encoding="utf-8") as f:
        return f.read()


@app.get("/api/dolar")
async def get_dolar():
    return {"valor": obtener_dolar()}


@app.post("/api/buscar_stream")
async def buscar_stream(
    archivo: UploadFile = File(...),
    posicion: int = Form(1),
    pct_menos: float = Form(1.0),
    margen: float = Form(20.0),
    iva: float = Form(10.5),
    tiendas_excluir: str = Form(""),
    dolar: float = Form(...),
    modo_calc: str = Form("posicion"),
    tiendas_calc: str = Form(""),
    imp_int_override: float = Form(-1),
):
    nombre_base = os.path.splitext(archivo.filename)[0]
    ext = os.path.splitext(archivo.filename)[1].lower() or ".xlsx"
    path_excel  = os.path.join(TEMP_DIR, "calcular" + ext)
    with open(path_excel, "wb") as f:
        shutil.copyfileobj(archivo.file, f)

    productos = leer_productos_calcular(path_excel)
    excluir   = [t.strip() for t in tiendas_excluir.split(",") if t.strip()]
    config    = {"posicion": posicion, "pct_menos": pct_menos, "margen": margen, "iva": iva, "imp_int": 0}
    tiendas_comparar = [t.strip().lower() for t in tiendas_calc.split(",") if t.strip()]

    async def generar():
        filas = []
        total = len(productos)

        for i, producto in enumerate(productos, 1):
            nombre       = producto["nombre"]
            imp_int_prod = imp_int_override if imp_int_override >= 0 else producto["imp_int"]

            msg = json.dumps({"tipo": "progreso", "actual": i, "total": total, "producto": nombre})
            yield f"data: {msg}\n\n"

            resultados = await scrape_hardgamers(nombre, max_results=50)
            ordenados  = filtrar_y_ordenar(resultados, excluir)

            if not ordenados:
                fila = {
                    "producto": nombre, "precio_venta": None,
                    "competidor_nombre": "SIN RESULTADOS", "competidor_tienda": "",
                    "posicion_real": None, "pct_menos": pct_menos,
                    "precio_compra_usd": None, "imp_int": imp_int_prod,
                }
            else:
                idx  = min(posicion - 1, len(ordenados) - 1)
                comp = ordenados[idx]
                pv   = calcular_precio_venta(comp["precio"], pct_menos)
                pusd = calcular_precio_compra_usd(pv, margen, iva, imp_int_prod, dolar)
                # Agrupar por tienda para modo tiendas
            por_tienda = {}
            for r in (ordenados or []):
                tn = r["tienda"].strip()
                if tn not in por_tienda:
                    por_tienda[tn] = {"nombre": r["nombre"], "precio": r["precio"],
                                      "tienda": tn, "url": r.get("url", "")}

            fila = {
                    "producto": nombre, "precio_venta": round(pv, 2),
                    "competidor_nombre": comp["nombre"], "competidor_tienda": comp["tienda"],
                    "competidor_url": comp.get("url", ""),
                    "posicion_real": ordenados.index(comp) + 1,
                    "pct_menos": pct_menos, "precio_compra_usd": round(pusd, 2),
                    "imp_int": imp_int_prod,
                    "por_tienda": por_tienda,
                    "tiendas_cmp": tiendas_comparar,
                    "modo_calc": modo_calc,
                    "iva": iva,
                }

            # Buscar en CompraGamer
            resultados_cg = await scrape_competidores(nombre)

            filas.append(fila)
            msg2 = json.dumps({"tipo": "resultado", "fila": fila})
            yield f"data: {msg2}\n\n"

            # Fila extra de CompraGamer
            if resultados_cg:
                cg_ordenados = sorted(resultados_cg, key=lambda x: x["precio"])
                cg = cg_ordenados[0]
                cg_pv = calcular_precio_venta(cg["precio"], pct_menos)
                cg_pusd = calcular_precio_compra_usd(cg_pv, margen, iva, imp_int_prod, dolar)
                fila_cg = {
                    "producto": nombre, "precio_venta": cg["precio"],
                    "competidor_nombre": cg["nombre"], "competidor_tienda": "CompraGamer",
                    "competidor_url": cg.get("url", ""),
                    "posicion_real": "-", "pct_menos": None,
                    "precio_compra_usd": round(cg_pusd, 2), "imp_int": imp_int_prod,
                    "es_compragamer": True,
                }
                filas.append(fila_cg)
                msg_cg = json.dumps({"tipo": "resultado", "fila": fila_cg})
                yield f"data: {msg_cg}\n\n"

        # Guardar estado completo para exportaciones dinámicas
        app.state.last_nombre_base  = nombre_base
        app.state.last_dolar        = dolar
        app.state.last_iva          = iva
        app.state.last_margen       = margen
        app.state.last_filas_hg     = [f for f in filas if not f.get("es_compragamer")]
        app.state.last_filas_cg     = [f for f in filas if f.get("es_compragamer")]
        app.state.last_filas_todas  = filas
        app.state.last_pct_menos    = pct_menos

        msg3 = json.dumps({"tipo": "fin", "dolar": dolar})
        yield f"data: {msg3}\n\n"

    return StreamingResponse(generar(), media_type="text/event-stream")


@app.post("/api/actualizar_margen")
async def actualizar_margen(data: dict):
    app.state.last_margen = data.get("margen", 20)
    app.state.last_iva    = data.get("iva", 10.5)
    return {"ok": True}


@app.get("/api/descargar/analisis")
async def descargar_analisis(margen: float = None):
    filas_hg = getattr(app.state, "last_filas_hg", [])
    filas_cg = getattr(app.state, "last_filas_cg", [])
    dolar    = getattr(app.state, "last_dolar", 0)
    iva      = getattr(app.state, "last_iva", 10.5)
    nombre_base = getattr(app.state, "last_nombre_base", "analisis")

    if margen is None:
        margen = getattr(app.state, "last_margen", 20)

    if not filas_hg and not filas_cg:
        return {"error": "No hay datos"}

    # Recalcular con margen actual
    filas_recalc = recalcular_filas(filas_hg + filas_cg, margen, iva, dolar)

    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(TEMP_DIR, f"{nombre_base}_analisis_{fecha}.xlsx")
    exportar_excel_analisis(filas_recalc, path, dolar, margen, iva)

    return FileResponse(path, filename=os.path.basename(path),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/descargar/vendedor")
async def descargar_vendedor(tipo: str = "hardgamers", margen: float = None):
    filas_hg = getattr(app.state, "last_filas_hg", [])
    filas_cg = getattr(app.state, "last_filas_cg", [])
    dolar    = getattr(app.state, "last_dolar", 0)
    iva      = getattr(app.state, "last_iva", 10.5)

    if margen is None:
        margen = getattr(app.state, "last_margen", 20)

    if tipo == "hardgamers":
        filas = filas_hg
        nombre_archivo = "vendedor_hardgamers.xlsx"
    elif tipo == "compragamer":
        filas = filas_cg
        nombre_archivo = "vendedor_compragamer.xlsx"
    else:
        filas = filas_hg + filas_cg
        nombre_archivo = "vendedor_ambos.xlsx"

    if not filas:
        return {"error": "No hay datos"}

    # Recalcular con margen actual
    filas_recalc = recalcular_filas(filas, margen, iva, dolar)

    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(TEMP_DIR, f"{fecha}_{nombre_archivo}")
    exportar_excel_vendedor_simple(filas_recalc, path)

    return FileResponse(path, filename=nombre_archivo,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def abrir_navegador():
    webbrowser.open("http://localhost:8000")


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 8000))
    is_local = port == 8000 and not os.environ.get("RENDER")
    print("=" * 50)
    print("  Herramienta de Precios")
    print(f"  http://localhost:{port}")
    print("  Ctrl+C para cerrar")
    print("=" * 50)
    if is_local:
        Timer(1.5, abrir_navegador).start()
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="warning")
